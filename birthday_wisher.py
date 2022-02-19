from datetime import datetime
from http import server
from openpyxl import workbook, load_workbook
import smtplib


def send_email(sender_email, sender_passwd, recipient_name, recipient_email, message):
    DOMAIN = 'smtp.gmail.com'
    PORT = 587

    server = smtplib.SMTP(DOMAIN, PORT)
    try:
        print(f"Sending your wishes to {recipient_name}...")
        server.starttls()
        server.login(sender_email, sender_passwd)
        server.sendmail(sender_email,
                        recipient_email, message)

        print('Mail sent!')

    except Exception as err:
        print(err)

    finally:
        server.quit()


def format_date(date):
    date_timestamp = date.split(' ')
    date = date_timestamp[0]
    return date.split('-')


my_email = input("Sender email: ")
my_passwd = input("Sender password: ")

wb = load_workbook('birthdays.xlsx')
ws = wb.active
max_col = ws.max_column
max_row = ws.max_row

present_day = datetime.now().day
present_month = datetime.now().month

for i in range(2, max_row + 1):
    reciever_name = ws.cell(row=i, column=1).value
    current_date = str(ws.cell(row=i, column=2).value)
    reciever_email = ws.cell(row=i, column=3).value
    birthday_message = f"Happy Birthday,{reciever_name}!!"

    year, month, day = format_date(current_date)
    is_birthday = int(day) == present_day and int(month) == present_month

    if is_birthday:
        send_email(my_email, my_passwd, reciever_name,
                   reciever_email, birthday_message)
